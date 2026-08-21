#!/usr/bin/env python3
"""
wikipedia_capacity_enrich.py
Enriches dim_venue.capacity by searching Wikipedia for venue infoboxes.
Uses the combined top-25-per-state CSV as input.

Usage:
    # Dry run — outputs review CSV, no DB writes
    python scripts/wikipedia_capacity_enrich.py --input exports/pipeline_reviews/top25_venues_by_state.csv

    # Live — writes confirmed matches to Supabase
    python scripts/wikipedia_capacity_enrich.py --input exports/pipeline_reviews/top25_venues_by_state.csv --live

    # Skip already-filled rows (default behaviour)
    # Force re-check even if capacity already set
    python scripts/wikipedia_capacity_enrich.py --input exports/pipeline_reviews/top25_venues_by_state.csv --force
"""

import argparse
import csv
import os
import re
import time
import json
from datetime import date
from difflib import SequenceMatcher

import requests
from dotenv import load_dotenv

load_dotenv()

SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_SERVICE_KEY")

WIKIPEDIA_API = "https://en.wikipedia.org/w/api.php"
REQUEST_DELAY = 3.0   # seconds between Wikipedia API calls (tunable via --delay)
RETRY_DELAYS  = [30, 60, 120]  # seconds to wait after successive 429s
USER_AGENT    = "GrooveprintCapacityBot/1.0 (grooveprint.app; data enrichment)"

# Venues to skip (parks, streets, unknown, festival grounds — capacity not meaningful)
SKIP_KEYWORDS = [
    "unknown venue", "park", "grounds", "speedway", "raceway",
    "street", "waterfront", "fremont street", "festival grounds",
    "lake", "field", "square", "plaza", "outdoor", "quidi vidi",
]

# Capacity sanity bounds for venues
MIN_CAPACITY = 100
MAX_CAPACITY = 200_000


# ---------------------------------------------------------------------------
# Wikipedia helpers
# ---------------------------------------------------------------------------

def _api_get(params: dict) -> dict | None:
    """GET the Wikipedia API with retry on 429. Returns None if all retries exhausted."""
    for attempt, wait in enumerate([0] + RETRY_DELAYS):
        if wait:
            print(f"         ⚠ 429 rate-limited — waiting {wait}s (attempt {attempt}/{len(RETRY_DELAYS)})")
            time.sleep(wait)
        try:
            resp = requests.get(WIKIPEDIA_API, params=params,
                                headers={"User-Agent": USER_AGENT}, timeout=10)
            if resp.status_code == 429:
                continue  # will use next wait from RETRY_DELAYS
            resp.raise_for_status()
            return resp.json()
        except requests.HTTPError:
            continue
    print(f"         ✗ Skipping — exhausted retries after repeated 429s")
    return None


def wikipedia_search(query: str, n: int = 3) -> list[dict]:
    """Return top-n Wikipedia search results for query (includes snippet for city check)."""
    data = _api_get({
        "action": "query",
        "list": "search",
        "srsearch": query,
        "format": "json",
        "srlimit": n,
        "srprop": "snippet|title",
    })
    if data is None:
        return []
    return data.get("query", {}).get("search", [])


def get_wikitext(title: str) -> str:
    """Fetch raw wikitext for a Wikipedia page title."""
    data = _api_get({
        "action": "query",
        "titles": title,
        "prop": "revisions",
        "rvprop": "content",
        "rvslots": "main",
        "format": "json",
    })
    if data is None:
        return ""
    pages = data.get("query", {}).get("pages", {})
    for page in pages.values():
        slots = page.get("revisions", [{}])[0].get("slots", {})
        return slots.get("main", {}).get("*", "") or ""
    return ""


def city_mentioned(city: str, title: str, snippet: str, wikitext: str = "") -> bool:
    """
    Check that the correct city appears in the match.
    Requires city in the article title OR in both the snippet and wikitext,
    to avoid coincidental city mentions in unrelated articles.
    """
    city_l = city.lower()
    clean_snippet = re.sub(r'<[^>]+>', '', snippet).lower()
    if city_l in title.lower():
        return True
    if city_l in clean_snippet and city_l in wikitext.lower():
        return True
    return False


def parse_capacity(wikitext: str) -> int | None:
    """Extract capacity integer from Wikipedia infobox wikitext."""
    # Match `| capacity = 12,500` or `| seating_capacity = 12500`
    pattern = re.compile(
        r'\|\s*(?:seating_)?capacity\s*=\s*([^\n\|\}]+)',
        re.IGNORECASE
    )
    for match in pattern.finditer(wikitext):
        raw = match.group(1).strip()
        # Strip wiki markup: refs, templates, bold/italic
        raw = re.sub(r'<ref[^>]*>.*?</ref>', '', raw, flags=re.DOTALL)
        raw = re.sub(r'\{\{[^}]+\}\}', '', raw)
        raw = re.sub(r"'''?", '', raw)
        # First number wins
        num = re.search(r'[\d,]+', raw)
        if num:
            val = int(num.group(0).replace(',', ''))
            if MIN_CAPACITY <= val <= MAX_CAPACITY:
                return val
    return None


def name_similarity(a: str, b: str) -> float:
    """Simple string similarity ratio."""
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def should_skip(venue_name: str) -> bool:
    n = venue_name.lower()
    return any(kw in n for kw in SKIP_KEYWORDS)


# ---------------------------------------------------------------------------
# Supabase write
# ---------------------------------------------------------------------------

def update_capacity(venue_id: int, capacity: int) -> bool:
    """Write capacity to dim_venue via Supabase REST."""
    if not SUPABASE_URL or not SUPABASE_KEY:
        raise RuntimeError("SUPABASE_URL / SUPABASE_SERVICE_ROLE_KEY not set")
    url = f"{SUPABASE_URL}/rest/v1/dim_venue"
    resp = requests.patch(
        url,
        params={"venue_id": f"eq.{venue_id}"},
        json={"capacity": capacity},
        headers={
            "apikey": SUPABASE_KEY,
            "Authorization": f"Bearer {SUPABASE_KEY}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        timeout=10,
    )
    return resp.status_code in (200, 204)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    global REQUEST_DELAY

    parser = argparse.ArgumentParser(description="Enrich dim_venue.capacity from Wikipedia")
    parser.add_argument("--input", required=True, help="Path to combined top-25 CSV/XLSX")
    parser.add_argument("--live", action="store_true", help="Write to Supabase (default: dry run)")
    parser.add_argument("--force", action="store_true", help="Re-check even if capacity already set")
    parser.add_argument("--state", help="Limit to a single state/province code (e.g. CA)")
    parser.add_argument("--delay", type=float, default=REQUEST_DELAY,
                        help=f"Seconds between API requests (default: {REQUEST_DELAY})")
    args = parser.parse_args()
    REQUEST_DELAY = args.delay

    # Read input — support both CSV and XLSX
    rows = []
    if args.input.endswith(".xlsx"):
        import openpyxl
        wb = openpyxl.load_workbook(args.input, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))
    else:
        with open(args.input, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.DictReader(f))

    # Filter
    if args.state:
        rows = [r for r in rows if str(r.get("state", "")).strip().upper() == args.state.upper()]


    today = date.today().isoformat()
    out_path = f"exports/pipeline_reviews/wikipedia_capacity_review_{today}.csv"
    os.makedirs("exports/pipeline_reviews", exist_ok=True)

    results = []
    skipped = queued = found = written = errors = 0

    print(f"{'='*66}")
    print(f"Wikipedia Capacity Enrichment  {'[LIVE]' if args.live else '[DRY RUN]'}")
    print(f"Input: {args.input}   Rows: {len(rows)}")
    print(f"{'='*66}\n")

    for row in rows:
        venue_id   = int(row["venue_id"])
        venue_name = str(row["venue_name"]).strip()
        city       = str(row["city"]).strip()
        state      = str(row["state"]).strip()
        existing   = row.get("capacity")

        # Skip if already filled (unless --force)
        if existing and str(existing).strip() not in ("", "NaN", "None", "null") and not args.force:
            skipped += 1
            continue

        if should_skip(venue_name):
            print(f"  SKIP  {venue_name} ({city}) — not a capacity venue")
            results.append({
                "venue_id": venue_id, "venue_name": venue_name,
                "city": city, "state": state,
                "capacity_found": "", "wikipedia_title": "",
                "similarity": "", "reason": "skipped_non_venue",
            })
            skipped += 1
            continue

        queued += 1
        query = f"{venue_name} {city} concert venue"
        print(f"  [{queued:3d}]  {venue_name[:40]:40s}  ({city}, {state})")

        try:
            time.sleep(REQUEST_DELAY)
            search_results = wikipedia_search(query)

            capacity = None
            matched_title = ""
            similarity = 0.0

            for result in search_results:
                title   = result["title"]
                snippet = result.get("snippet", "")
                sim = name_similarity(venue_name, title)

                # Name must be plausible
                if sim < 0.35 and venue_name.lower() not in title.lower():
                    continue

                time.sleep(REQUEST_DELAY)
                wikitext = get_wikitext(title)

                # City must appear in title OR in both snippet and wikitext
                if not city_mentioned(city, title, snippet, wikitext):
                    continue

                cap = parse_capacity(wikitext)

                if cap:
                    capacity = cap
                    matched_title = title
                    similarity = round(sim, 3)
                    break  # take first good hit

            if capacity:
                found += 1
                flag = "LOW_CONFIDENCE" if similarity < 0.6 else "ok"
                print(f"         → {capacity:,}  [{matched_title}]  sim={similarity}  {flag}")
                results.append({
                    "venue_id": venue_id, "venue_name": venue_name,
                    "city": city, "state": state,
                    "capacity_found": capacity, "wikipedia_title": matched_title,
                    "similarity": similarity, "reason": flag,
                })
                if args.live and flag != "LOW_CONFIDENCE":
                    ok = update_capacity(venue_id, capacity)
                    if ok:
                        written += 1
                    else:
                        errors += 1
                        print(f"         ✗ Supabase write failed")
            else:
                print(f"         → no capacity found")
                results.append({
                    "venue_id": venue_id, "venue_name": venue_name,
                    "city": city, "state": state,
                    "capacity_found": "", "wikipedia_title": matched_title or "",
                    "similarity": similarity or "", "reason": "no_capacity_found",
                })

        except Exception as e:
            print(f"         ✗ ERROR: {e}")
            results.append({
                "venue_id": venue_id, "venue_name": venue_name,
                "city": city, "state": state,
                "capacity_found": "", "wikipedia_title": "",
                "similarity": "", "reason": f"error: {e}",
            })
            errors += 1

    # Write review CSV
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        fieldnames = ["venue_id","venue_name","city","state",
                      "capacity_found","wikipedia_title","similarity","reason"]
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    print(f"\n{'='*66}")
    print(f"Done.  Queued: {queued}  Found: {found}  Written: {written}  Errors: {errors}  Skipped: {skipped}")
    print(f"Review CSV: {out_path}")
    print(f"{'='*66}")


if __name__ == "__main__":
    main()
